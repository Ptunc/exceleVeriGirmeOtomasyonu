from openpyxl import Workbook, load_workbook

#hangi sütun aralığını değiştireceğinin verisini alan ve o sütun aralığında X gördüğü yeri istenen formülle değiştiren kod
def anaFonk(sutunBaslangic,sutunBitis):

    for satir in range(5, ws.max_row):
        for sutun in range(sutunBaslangic,sutunBitis):
            if(str(ws.cell(satir,sutun).value) == "X"):
                ws.cell(satir, sutun).value = "=+C" + str(satir)


#belirli bir satır-sütun aralığında X gördüğü yeri değişkenleri birbirine ekleyerek istenen formülle değiştiren kod
def ozelCarpimFonk(satirBaslangic, satirBitis,sutunBaslangic,sutunBitis,carpim):

    for satir in range(satirBaslangic, satirBitis):
        for sutun in range(sutunBaslangic,sutunBitis):
            if(str(ws.cell(satir,sutun).value) == "X"):
                ws.cell(satir, sutun).value = "=+D"+ str(satir) + "*" + carpim + "-E" + str(satir)



wb = load_workbook("deneme.xlsx")
ws = wb.active

#fonksiyonlara gerekli değişkenleri vererek çağırma
anaFonk(6,18)
anaFonk(29,70)
ozelCarpimFonk(5, 120, 19, 22, "4.03")
ozelCarpimFonk(120, 160, 19, 22, "5.53")
ozelCarpimFonk(5, 120, 22, 23, "3.6")
ozelCarpimFonk(120, 160, 22, 23, "5.03")
ozelCarpimFonk(5, 120, 23, 29, "4.03")
ozelCarpimFonk(120, 160, 23, 29, "5.53")


wb.save("denemenindenemesi.xlsx")



